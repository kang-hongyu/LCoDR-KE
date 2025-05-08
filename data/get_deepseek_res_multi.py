#coding:utf8
import csv
import time
import sys
import json
import pandas as pd
import openpyxl
from openpyxl import Workbook
import os
import threading
from queue import Queue
from openai import OpenAI

client = OpenAI(api_key="sk-xxx", base_url="https://api.deepseek.com")

# 配置参数
NUM_WORKERS = 10  # 工作线程数量
MAX_TASKS = 2000  # 最大处理任务数
csv.field_size_limit(65536)

# 全局队列
task_queue = Queue(maxsize=100)  # 控制内存使用
result_queue = Queue()

# 输出文件配置
EXCEL_OUTPUT = "output_deepseek_r1_else2.xlsx"
JSONL_OUTPUT = "output_deepseek_r1_else2.jsonl"

def get_chatgpt_response(question):
    conversation = []
    prompt = '''You are a biomedical knowledge graph construction assistant. Perform document analysis and relationship extraction following these strict protocols:

# Processing Workflow
1. **Schema Identification**
   - Identification of Entity and Relationship Types:
   • Identify candidate entities (12 predefined types) 
   • Detect contextual relationships (26 predefined classes)
   • Map findings to the following predefined schema:

2. **Entity Recognition** 
   Types with Definitions:
   1). Anatomy : Structures or functional parts of the body systems (e.g., organs, tissues, or cellular components). 
Key distinction: Explicitly excludes pathological states (e.g., "inflamed liver" would belong to disease/symptom).  
Example: liver, hippocampus, T lymphocytes.

2). Biomarker: A measurable biological indicator (e.g., molecular, genetic, or biochemical) used to assess physiological/pathological states or responses to interventions, typically acting as an indirect indicator (not the direct target of a drug). 
Example: hemoglobin A1c (HbA1c), C-reactive protein.

3). Complication: A secondary medical condition arising as a direct consequence of a primary disease or medical intervention (e.g., sepsis in burn patients).*  
Key distinction: Must involve causality to a pre-existing condition.  
Example: diabetic neuropathy in diabetes mellitus.

4). Disease: A specific pathological disorder with characteristic signs/symptoms, affecting structure or function, not inherently implying a causal relationship to another condition 
Example: Alzheimer's disease, rheumatoid arthritis.

5). Drug: A chemical substance with therapeutic/diagnostic properties, including generic names, brand names, and investigational compounds. Excludes non-pharmacological interventions (e.g., surgery).  
Example: aspirin, Bevacizumab (Avastin®).

6). Gene: A functional unit of heredity represented by standardized symbols or names, distinct from gene products (e.g., proteins) classified under "Target" 
Example: BRCA1, TP53.

7). Side Effect: An unintended physiological response **directly linked to pharmacological actionof a drug at normal doses (e.g., nausea caused by chemotherapy). 
Key distinction: Requires explicit association with drug exposure.  

8). Symptom: A patient-perceived subjective manifestation of a disease or condition, independent of drug exposure (e.g., fatigue in cancer). 
Example: fatigue, chest pain.

9). Target: A molecular entity (protein, enzyme, receptor, etc.) directly modulated by a drug to exert therapeutic effects (e.g., ACE2 receptor).  
Key distinction: Functional interaction with a drug is required.  

10). Test: A diagnostic procedure or assay to detect/measure biological markers, disease states, or treatment responses, excluding the biomarkers themselves  
Example: MRI scan, ELISA assay.

11). Treatment: A clinical intervention (pharmacological, surgical, or behavioral) intended to prevent/manage diseases, including non-pharmacological approaches (e.g., radiotherapy). 
Key distinction: Broader scope than "Drug".  

   Annotation Principles:
   (1) Non-overlapping: Single text span → single entity type
   (2) Non-nesting: No embedded entities within spans
   (3) Minimal punctuation: Exclude non-essential conjunctions/punctuation

3. **Relation Extraction** 
   **Relationship Definitions**  
1). complication_of | Complication ↔ Disease
Definition:
A complication is a secondary medical condition directly arising from a primary disease or its treatment (e.g., surgery, drug therapy). This relationship implies bidirectional causality: the disease initiates the complication, and the complication cannot exist independently of the disease.
Key distinctions:
•	Must involve a causal link (e.g., diabetic retinopathy is caused by diabetes, not coincidental).
•	Excludes unrelated coexisting diseases (e.g., hypertension and asthma in the same patient).
Examples:
•	"Diabetic neuropathy" complication_of "Diabetes mellitus" (explicit).
•	"Sepsis" complication_of "Burn injury" (implicit: burns → infection → sepsis).
________________________________________
2). increases_expression_of | Drug ↔ Gene
Definition:
A drug pharmacologically enhances the transcription or translation activity of a gene, leading to measurable increases in its mRNA or protein levels. 
Key distinctions:
•	Gene expression changes must be drug-specific (not general cellular stress responses).
Examples:
•	"Tamoxifen" increases_expression_of "ESR1" (estrogen receptor gene).
•	"Metformin" increases_expression_of "AMPK" (via metabolic pathway activation).
________________________________________
3). is_biomarker_of | Biomarker ↔ Disease
Definition:
A biomarker (e.g., protein, metabolite) is objectively measured to indicate the presence, severity, or progression of a disease. The biomarker has diagnostic, prognostic, or therapeutic monitoring utility.
Key distinctions:
•	Biomarkers are indicators, not therapeutic targets (contrast with is_target_of).
Examples:
•	"CA-125" is_biomarker_of "Ovarian cancer" (diagnostic).
•	"Tau protein" is_biomarker_of "Alzheimer’s disease" (progression monitoring).
________________________________________
4). is_located_in | Disease ↔ Anatomy
Definition:
A disease primarily manifests in or affects a specific anatomical structure (organ, tissue, or cell type). The anatomical site is pathologically or functionally central to the disease.
Key distinctions:
•	Anatomical specificity is required (e.g., "lung" for pneumonia, not "respiratory system").
•	Excludes systemic diseases without localized pathology (e.g., sepsis).
Examples:
•	"Hepatitis B" is_located_in "Liver" (viral replication in hepatocytes).
•	"Glioblastoma" is_located_in "Brain" (primary tumor site).
________________________________________
5). is_side_effect_of | Side Effect ↔ Drug
Definition:
A side effect is an unintended physiological reaction directly attributable to the pharmacological action of a drug at therapeutic doses. The reaction must occur in a statistically significant patient population.
Key distinctions:
•	Must exclude symptoms caused by the disease itself (e.g., chemotherapy-induced nausea vs. cancer-related fatigue).
•	Requires dose-dependency (higher dose → higher risk).
Examples:
•	"Hemorrhage" is_side_effect_of "Warfarin" (anticoagulant effect).
•	"Insomnia" is_side_effect_of "Dexamethasone" (CNS stimulation).
________________________________________
6). is_symptom_of | Symptom ↔ Disease
Definition:
A symptom is a patient-reported or clinically observed manifestation (subjective or objective) that is pathognomonic or commonly associated with a disease. Symptoms arise from the disease’s pathophysiology.
Key distinctions:
•	Symptoms are disease-specific (e.g., "jaundice" in hepatitis).
•	Excludes nonspecific complaints (e.g., "fatigue" without disease context).
Examples:
•	"Dyspnea" is_symptom_of "Heart failure" (fluid accumulation in lungs).
•	"Hematuria" is_symptom_of "Bladder cancer" (tissue invasion).
________________________________________
7). is_target_of | Target ↔ Drug
Definition:
A target is a biomolecule (protein, receptor, enzyme) that directly interacts with a drug to mediate its therapeutic effect. The interaction is stoichiometric and mechanistically validated (e.g., binding assays, crystallography).
Key distinctions:
•	Targets are functional entities (e.g., HER2 protein, not the HER2 gene).
Examples:
•	"ACE2 receptor" is_target_of "Losartan" (angiotensin receptor blocker).
•	"COX-1 enzyme" is_target_of "Aspirin" (irreversible acetylation).
________________________________________
8). treat | Drug ↔ Disease
Definition:
A drug is clinically used to alleviate, manage, or cure a disease through a mechanism of action supported by regulatory approval or evidence-based guidelines. The relationship implies therapeutic efficacy.
Key distinctions:
•	Requires clinical relevance (e.g., "insulin treats diabetes," not "vitamin C treats common cold").
Examples:
•	"Penicillin" treat "Streptococcal pharyngitis" (antibacterial action).
•	"Sertraline" treat "Major depressive disorder" (SSRI mechanism).
________________________________________
9). is_examination_for | Test ↔ Disease
Definition:
A test is a standardized diagnostic or monitoring procedure (imaging, lab assay) used to confirm, stage, or track the progression of a disease. The test must have established clinical utility.
Key distinctions:
•	Tests are actions/tools, not biomarkers (e.g., "MRI" is a test; "elevated CRP" is a biomarker).
•	Excludes research-only or experimental assays.
Examples:
•	"Colonoscopy" is_examination_for "Colorectal cancer" (diagnostic).
•	"Electrocardiogram (ECG)" is_examination_for "Myocardial infarction" (diagnostic).

   Annotation Principles:
   (1) Intra-sentence priority: Prefer relations within single sentences
   (2) Unidirectionality: Maintain only one directional relation per entity pair
   (3) Schema compliance: Use only predefined relationship types

# Output Specifications
• Strict JSON format with two root keys: Entities, Relationships
• Entity preservation: Maintain original text case and formatting
• Relationship format: [source_entity, source_type, relationship, target_entity, target_type]
• Please makesure the entity and relationship is from the input text
• No null/empty values or placeholder text
• No explanatory content

# Input Text
'''+question+'''

# Output Example
{
    "Entities": {
        "Hippocampal formation": "anatomy",
        "Tau protein phosphorylation": "biomarker",
        "Neuropsychiatric symptoms": "complication",
        "Alzheimer's disease": "disease",
        "Lecanemab": "drug",
        "APOE ε4 allele": "gene",
        "Amyloid-related imaging abnormalities": "side effect",
        "Cognitive decline": "symptom",
        "Amyloid-β protofibrils": "target",
        "PET scan": "test",
        "Anti-amyloid immunotherapy": "treatment"
    },
    "Relationships": [
        ["Alzheimer's disease", "disease", "have_complication", "Neuropsychiatric symptoms", "complication"],
        ["Lecanemab", "drug", "is_target_drug_of", "Amyloid-β protofibrils", "target"],
        ["Tau protein phosphorylation", "biomarker", "is_biomarker_of", "Alzheimer's disease", "disease"],
        ["Cognitive decline", "symptom", "is_symptom_of", "Alzheimer's disease", "disease"],
        ["APOE ε4 allele", "gene", "increases_expression_of", "Amyloid-β protofibrils", "target"]
    ]
}'''
    conversation.append({"role":"user","content":prompt})
    retry = 0
    res = ""
    reasoning_content = ""
    while retry < 10:
        try:
            response = client.chat.completions.create(
                model="deepseek-reasoner",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant"},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=8192,
                temperature=0.01,
                stream=False
            )
            res = response.choices[0].message.content
            reasoning_content = response.choices[0].message.reasoning_content
            #print(response)
            break
        except Exception as e:
            retry += 1
            time.sleep(30)
            print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
            print("get deepseek response error! retry!")
            print(e)
            print(question)
    return res,reasoning_content,conversation

def extract_json_strings(content):
    """提取所有最外层JSON字符串的辅助函数"""
    json_objects = []
    start = 0
    while True:
        start = content.find('{', start)
        if start == -1:
            break
        
        end = start
        brace_count = 0
        while end < len(content):
            if content[end] == '{':
                brace_count += 1
            elif content[end] == '}':
                brace_count -= 1
            
            if brace_count == 0:  # 找到一个完整的最外层 JSON
                break
            end += 1
        
        if brace_count == 0:  # 确保找到了匹配的括号
            json_str = content[start:end + 1]
            json_str = json_str.replace('\r\n', '').replace('  ', ' ').replace('\n', '')
            json_str = json_str.replace('{\n', '{').replace(', ],', ' ],')
            
            try:
                json_obj = json.loads(json_str, strict=False)
                json_objects.append(json_obj)
            except json.JSONDecodeError:
                fixed_json = try_fix_json(json_str)
                json_objects.append(json.loads(fixed_json))
        
        start = end + 1  # 移动到下一个可能的 JSON 开始位置

    return json_objects[0] if json_objects else None

def try_fix_json(json_string):
    bracket_map = {'}': '{', ']': '[', '{': '}', '[': ']'}
    stack = []
    fixed_string = ''

    for char in json_string:
        if char in ['{', '[']:
            stack.append(char)
            fixed_string += char
        elif char in ['}', ']']:
            if stack and stack[-1] == bracket_map[char]:
                stack.pop()
                fixed_string += char
            else:
                continue
        else:
            fixed_string += char

    while stack:
        fixed_string += bracket_map[stack.pop()]
    while fixed_string.startswith("{{") and fixed_string.endswith("}}"):
        fixed_string = fixed_string[1:-1]

    return fixed_string
def parse_response(response_json, row_id, row_content):
    try:
        response_dict = extract_json_strings(response_json)
    except json.JSONDecodeError:
        print("JSON 解析错误:", response_json)
        return None,None,None

    entities = []
    relationships = []

    # 解析Entities
    if "Entities" in response_dict:
        for name, entity_type in response_dict["Entities"].items():
            entities.append({
                "id": row_id,
                "content": row_content,
                "entity_type": entity_type,
                "name": name
            })

    # 解析Relationships
    if "Relationships" in response_dict:
        for relationship in response_dict["Relationships"]:
            relationships.append({
                "id": row_id,
                "content": row_content,
                "entity_type_1": relationship[1],
                "entity_name_1": relationship[0],
                "relationship": relationship[2],
                "entity_type_2": relationship[4],
                "entity_name_2": relationship[3]
            })

    return entities, relationships

def append_to_excel(entities, relationships, output_file):
    """增量追加数据到Excel文件"""
    # 获取或创建所有必需的sheet
    sheets = {
        "Entities": ["id", "content", "entity type", "name"],
        "Relationships": ["id", "content", "entity type 1", "entity name 1", 
                        "relationship", "entity type 2", "entity name 2"]
    }
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(sheets[sheet_name])
    else:
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认sheet
        
        # 创建并初始化所有sheet
        entity_sheet = wb.create_sheet("Entities")
        entity_sheet.append(sheets["Entities"])
        
        relationship_sheet = wb.create_sheet("Relationships")
        relationship_sheet.append(sheets["Relationships"])

    # 追加数据
    for entity in entities:
        wb["Entities"].append([
            entity["id"], entity["content"],
            entity["entity_type"], entity["name"]
        ])
    
    for rel in relationships:
        wb["Relationships"].append([
            rel["id"], rel["content"],
            rel["entity_type_1"], rel["entity_name_1"],
            rel["relationship"], rel["entity_type_2"],
            rel["entity_name_2"]
        ])

    wb.save(output_file)

def worker():
    while True:
        task = task_queue.get()
        if task is None:  # 终止信号
            task_queue.task_done()
            break
        
        row_id, row_content = task
        try:
            if len(row_content) < 30:
                print(f"Skipping short content: ID {row_id}")
                continue

            # 获取API响应
            response, reasoner, conversation = get_chatgpt_response(row_content)
            
            # 清理响应内容
            if response and not response.startswith("{"):
                response = response[response.find("{"):response.rfind("}")+1]
            
            # 解析结果
            entities, relationships = parse_response(response, row_id, row_content)
            
            if entities or relationships:
                result_queue.put((
                    entities,
                    relationships,
                    {
                        "id": row_id,
                        "conversation": conversation,
                        "response": response,
                        "reasoner": reasoner
                    }
                ))
            else:
                print(f"Empty result for ID {row_id}")
                
        except Exception as e:
            print(f"Error processing ID {row_id}: {str(e)}")
        finally:
            task_queue.task_done()

def writer():
    """专用的写入线程"""
    while True:
        result = result_queue.get()
        if result is None:  # 终止信号
            result_queue.task_done()
            break
        
        try:
            entities, relationships, meta_data = result
            
            # 写入Excel
            append_to_excel(entities, relationships, EXCEL_OUTPUT)
            
            # 写入JSONL
            with open(JSONL_OUTPUT, "a", encoding="utf-8") as f:
                json.dump(meta_data, f, ensure_ascii=False)
                f.write("\n")
                
        except Exception as e:
            print(f"写入错误: {str(e)}")
        finally:
            result_queue.task_done()

def main():
    # 初始化环境
    #if os.path.exists(EXCEL_OUTPUT):
    #    os.remove(EXCEL_OUTPUT)
    #if os.path.exists(JSONL_OUTPUT):
    #    os.remove(JSONL_OUTPUT)
    time.sleep(7200)
    # 读取数据
    df = pd.read_excel("hymax_literature_khy_20230612_else.xlsx")
    ready_data = {}
    with open("output_deepseek_r1_else.jsonl", "r", encoding='utf8') as f:
        for line in f:
            data = json.loads(line)
            ready_data[data["id"]] = True

    # 创建并启动工作线程
    workers = []
    for _ in range(NUM_WORKERS):
        t = threading.Thread(target=worker)
        t.start()
        workers.append(t)

    # 启动写入线程
    writer_thread = threading.Thread(target=writer)
    writer_thread.start()

    # 填充任务队列
    count = 0
    for index, row in df.iterrows():
        if index <= 8000:
            continue
        if count >= MAX_TASKS:
            break
        if row[0] in ready_data:
            continue
        if len(row[2]) >= 30:  # 预处理过滤短内容
            task_queue.put((row[0], row[2]))
            count += 1

    # 发送终止信号给工作线程
    for _ in range(NUM_WORKERS):
        task_queue.put(None)

    # 等待任务完成
    task_queue.join()

    # 发送终止信号给写入线程
    result_queue.put(None)

    # 等待所有线程结束
    for t in workers:
        t.join()
    writer_thread.join()

if __name__ == "__main__":
    main()